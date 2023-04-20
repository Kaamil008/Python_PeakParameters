#!/usr/bin/env python
# coding: utf-8

# In[6]:


import numpy as np
import matplotlib.pyplot as plt
from scipy import interpolate
from scipy.signal import *
import pandas as pd
from openpyxl import Workbook, load_workbook


# In[7]:


def peak_width(x,y,Max_peak):
    xnew = np.linspace(min(x),max(x),10000)
    f = interpolate.interp1d(x, y)
    ynew = f(xnew)
    target = Max_peak*0.5
    print(target)
    b=0
    a=0
    ay=0
    by=0
    for i in range(len(xnew)):
        if ynew[i]-target>0:
            a=xnew[i]
            j=i
            ay=ynew[i]
            break
    for i in range(j,len(xnew)):
        if ynew[i]-target<0:
            b=xnew[i]
            by=ynew[i]
            break
    return a,ay,b,by,b-a


# In[8]:


path=r'C:\Users\kaamil.kaleem\OneDrive - Vibracoustic\Desktop\test_abc.xlsx'
#path_2=r'C:\Users\kaamil.kaleem\OneDrive - Vibracoustic\Desktop\test.xlsx'
wb = load_workbook(path)
ws_all = wb.sheetnames
for i in range(32):
    nor = pd.read_excel(path, sheet_name=i)
    norm_fre = nor['Norm Frq']
    norm_dynstiffness = nor['Norm Dyn']
    x=np.array(norm_fre)
    y=np.array(norm_dynstiffness)
    peaks =find_peaks(y)
    print(peaks)
    height = y[peaks[0]]
    print(height)
    peak_pos = x[peaks[0]]
    print(peak_pos)
    Max_peak = height[0]
    width=peak_width(x,y,Max_peak)
    fig =plt.figure()
    ax = fig.subplots()
    ax.plot(x,y)
    ax.scatter(peak_pos, height, color='r', s=15, marker='D', label='Peak')
    pwx=[width[0], width[2]]
    pwy=[width[1], width[3]]
    ax.plot(pwx,pwy,'bo', markersize=2,ls='--')
    ax.plot()
    plt.legend()
    plt.grid()
    plt.show()
    Frq_peak = peak_pos[0]
    print(Max_peak, width[4], Frq_peak)
    ws = ws_all[i]
    k=wb[ws]
    k['I1'].value=Max_peak
    k['I2'].value=width[4]
    k['I3'].value=Frq_peak
    k['J2']=''
    k['L2']=''
    k['N2']=''
wb.save(path)


# In[11]:


path=r'C:\Users\kaamil.kaleem\OneDrive - Vibracoustic\Desktop\test_abc.xlsx'
path_2=r'C:\Users\kaamil.kaleem\OneDrive - Vibracoustic\Desktop\Peak_Parameters.txt'
wb = load_workbook(path)
wb_all=wb.sheetnames
f = open(path_2,'a')
f.write(f"a,b,c,max_peak,width_peak,fre_peak\n")
for i in range(len(wb_all)):
    ws=wb_all[i]
    k=wb[ws]
    a=k['B2'].value
    b=k['B3'].value
    c=k['B4'].value
    max_peak=k['I1'].value
    width_peak=k['I2'].value
    fre_peak=k['I3'].value
    f.write(f"{a},{b},{c},{max_peak},{width_peak},{fre_peak}\n")
f.close()
wb.save(path)
    


# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:




